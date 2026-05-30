import io
from datetime import datetime
from flask import render_template, request, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app import db
from app.models.admin_log import AdminLog
from app.routes.admin import admin_bp, admin_required
from app.utils.op_log import cleanup_old_logs, log_operation


@admin_bp.route('/op-logs')
@login_required
@admin_required
def op_logs():
    """操作日志列表"""
    page = request.args.get('page', 1, type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    admin_name = request.args.get('admin_name', '').strip()
    action_type = request.args.get('action_type')
    target_type = request.args.get('target_type')
    result = request.args.get('result')

    query = AdminLog.query

    if start_date:
        try:
            sd = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(AdminLog.created_at >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(AdminLog.created_at <= ed)
        except ValueError:
            pass
    if admin_name:
        query = query.filter(AdminLog.admin_name.contains(admin_name))
    if action_type:
        query = query.filter_by(action_type=action_type)
    if target_type:
        query = query.filter_by(target_type=target_type)
    if result:
        query = query.filter_by(result=result)

    pagination = query.order_by(AdminLog.created_at.desc()).paginate(page=page, per_page=30)

    # 操作类型选项
    action_types = [
        ('create', '创建'), ('update', '编辑'), ('delete', '删除'),
        ('export', '导出'), ('import', '导入'), ('config_change', '配置变更'),
        ('batch_operation', '批量操作')
    ]

    # 对象类型选项
    target_types = [
        ('user', '用户'), ('subject', '学科'), ('chapter', '章节'),
        ('question', '题目'), ('setting', '设置'), ('announcement', '公告'),
        ('backup', '备份'), ('dictionary', '字典'), ('rank_tier', '段位'),
        ('leaderboard', '排行榜'), ('security', '安全')
    ]

    # 结果选项
    result_options = [('success', '成功'), ('failure', '失败')]

    breadcrumb = [
        {'label': '系统管理'},
        {'label': '操作日志'}
    ]

    return render_template('admin/op_logs.html',
                          pagination=pagination,
                          action_types=action_types,
                          target_types=target_types,
                          result_options=result_options,
                          breadcrumb=breadcrumb,
                          start_date=start_date,
                          end_date=end_date,
                          admin_name=admin_name,
                          action_type=action_type,
                          target_type=target_type,
                          result=result)


@admin_bp.route('/op-logs/export')
@login_required
@admin_required
def export_op_logs():
    """导出操作日志为Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    admin_name = request.args.get('admin_name', '').strip()
    action_type = request.args.get('action_type')
    target_type = request.args.get('target_type')
    result = request.args.get('result')

    query = AdminLog.query

    if start_date:
        try:
            sd = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(AdminLog.created_at >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.strptime(end_date, '%Y-%m-%d')
            query = query.filter(AdminLog.created_at <= ed)
        except ValueError:
            pass
    if admin_name:
        query = query.filter(AdminLog.admin_name.contains(admin_name))
    if action_type:
        query = query.filter_by(action_type=action_type)
    if target_type:
        query = query.filter_by(target_type=target_type)
    if result:
        query = query.filter_by(result=result)

    logs = query.order_by(AdminLog.created_at.desc()).limit(10000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "操作日志"

    headers = ['时间', '操作人', '操作类型', '对象类型', '对象ID', '对象名称', '结果', 'IP地址', '详情']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '')
        ws.cell(row=row_idx, column=2, value=log.admin_name or '')
        ws.cell(row=row_idx, column=3, value=log.action_type or '')
        ws.cell(row=row_idx, column=4, value=log.target_type or '')
        ws.cell(row=row_idx, column=5, value=log.target_id or '')
        ws.cell(row=row_idx, column=6, value=log.target_name or '')
        ws.cell(row=row_idx, column=7, value='成功' if log.result == 'success' else '失败')
        ws.cell(row=row_idx, column=8, value=log.ip_address or '')
        ws.cell(row=row_idx, column=9, value=log.detail or '')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_operation('export', 'audit_log', detail=f'导出 {len(logs)} 条操作日志')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='操作日志.xlsx'
    )

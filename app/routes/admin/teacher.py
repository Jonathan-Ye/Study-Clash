"""教师专属路由 - 学生导入功能"""
import io
import re
from datetime import datetime
from flask import render_template, request, redirect, url_for, flash, send_file, session
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app import db
from app.models import User, DictionaryItem
from app.routes.admin import admin_bp, role_required, teacher_permission_required
from app.utils.op_log import log_operation

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
MAX_IMPORT_ROWS = 500


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@admin_bp.route('/teacher/student-import')
@login_required
@role_required('teacher')
@teacher_permission_required('can_import_students')
def teacher_student_import():
    """教师学生导入页面"""
    breadcrumb = [
        {'label': '学生管理'},
        {'label': '学生导入'}
    ]
    return render_template('admin/teacher_student_import.html', breadcrumb=breadcrumb)


@admin_bp.route('/teacher/student-import/validate', methods=['POST'])
@login_required
@role_required('teacher')
@teacher_permission_required('can_import_students')
def teacher_student_import_validate():
    """校验导入文件"""
    if 'file' not in request.files:
        flash('未选择文件', 'error')
        return redirect(url_for('admin.teacher_student_import'))

    file = request.files['file']
    if file.filename == '':
        flash('未选择文件', 'error')
        return redirect(url_for('admin.teacher_student_import'))

    if not allowed_file(file.filename):
        flash('不支持的文件格式，请上传 .xlsx, .xls 或 .csv 文件', 'error')
        return redirect(url_for('admin.teacher_student_import'))

    try:
        import pandas as pd

        # 读取文件内容
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file, encoding='utf-8-sig')
        else:
            df = pd.read_excel(file)

        # 数据量限制
        if len(df) > MAX_IMPORT_ROWS:
            flash(f'数据量过大！单次最多导入 {MAX_IMPORT_ROWS} 条记录，当前有 {len(df)} 条', 'error')
            return redirect(url_for('admin.teacher_student_import'))

        # 标准化列名
        column_mapping = {
            'username': ['username', '用户名', '账号', '登录名'],
            'real_name': ['real_name', '姓名', '真实姓名', 'name'],
            'password': ['password', '密码', '登陆密码'],
            'email': ['email', '邮箱', '电子邮件', 'Email'],
            'nickname': ['nickname', '昵称', '别名'],
            'student_id': ['student_id', '学号'],
            'phone': ['phone', '手机', '手机号码', '电话'],
            'school': ['school', '学校'],
            'grade': ['grade', '年级'],
            'major': ['major', '专业', '文理科'],
            'class_name': ['class_name', '班级', 'ClassName']
        }

        reverse_mapping = {}
        for standard_col, variants in column_mapping.items():
            for variant in variants:
                reverse_mapping[variant.lower()] = standard_col

        new_columns = {}
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if col_lower in reverse_mapping:
                new_columns[col] = reverse_mapping[col_lower]
            else:
                new_columns[col] = col_lower

        df.rename(columns=new_columns, inplace=True)

        valid_users = []
        invalid_users = []
        total_rows = len(df)

        for idx, row in df.iterrows():
            row_num = idx + 2

            user_data = {
                'row': row_num,
                'username': str(row.get('username', '')).strip() if pd.notna(row.get('username')) else '',
                'real_name': str(row.get('real_name', '')).strip() if pd.notna(row.get('real_name')) else '',
                'password': str(row.get('password', '')).strip() if pd.notna(row.get('password')) else '',
                'email': str(row.get('email', '')).strip() if pd.notna(row.get('email')) else '',
                'nickname': str(row.get('nickname', '')).strip() if pd.notna(row.get('nickname')) else '',
                'student_id': str(row.get('student_id', '')).strip() if pd.notna(row.get('student_id')) else '',
                'phone': str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else '',
                'school': str(row.get('school', '')).strip() if pd.notna(row.get('school')) else '',
                'grade': str(row.get('grade', '')).strip() if pd.notna(row.get('grade')) else '',
                'major': str(row.get('major', '')).strip() if pd.notna(row.get('major')) else '',
                'class_name': str(row.get('class_name', '')).strip() if pd.notna(row.get('class_name')) else ''
            }

            errors = []

            if not user_data['username']:
                errors.append('用户名为空')
            elif len(user_data['username']) < 3:
                errors.append('用户名太短（至少3个字符）')
            elif not re.match(r'^[a-zA-Z0-9_]+$', user_data['username']):
                errors.append('用户名格式不正确')

            if not user_data['real_name']:
                errors.append('姓名为空')

            if not user_data['password']:
                errors.append('密码为空')
            elif len(user_data['password']) < 6:
                errors.append('密码太短（至少6个字符）')

            if not user_data['student_id']:
                errors.append('学号为空')
            elif not re.match(r'^[0-9]+$', user_data['student_id']):
                errors.append('学号必须是纯数字')

            if user_data['username'] and User.query.filter_by(username=user_data['username']).first():
                errors.append(f'用户名 "{user_data["username"]}" 已存在')

            if user_data['email']:
                if not re.match(r'^[^@]+@[^@]+\.[^@]+$', user_data['email']):
                    errors.append('邮箱格式不正确')
                elif User.query.filter_by(email=user_data['email']).first():
                    errors.append(f'邮箱 "{user_data["email"]}" 已存在')

            if errors:
                invalid_users.append({
                    'row': row_num,
                    'real_name': user_data['real_name'] or '-',
                    'student_id': user_data['student_id'] or '-',
                    'errors': errors
                })
            else:
                if not user_data['email']:
                    user_data['email'] = f"{user_data['username']}@studyclash.local"
                if not user_data['nickname']:
                    user_data['nickname'] = user_data['real_name']
                valid_users.append(user_data)

        session['teacher_import_data'] = {
            'valid_users': valid_users,
            'invalid_users': invalid_users,
            'total': total_rows
        }

        breadcrumb = [
            {'label': '学生管理'},
            {'label': '学生导入'},
            {'label': '导入预览'}
        ]

        return render_template(
            'admin/teacher_student_import.html',
            breadcrumb=breadcrumb,
            step='preview',
            total=total_rows,
            valid_users=valid_users,
            invalid_users=invalid_users
        )

    except Exception as e:
        flash(f'解析文件失败: {str(e)}', 'error')
        return redirect(url_for('admin.teacher_student_import'))


@admin_bp.route('/teacher/student-import/confirm', methods=['POST'])
@login_required
@role_required('teacher')
@teacher_permission_required('can_import_students')
def teacher_student_import_confirm():
    """确认导入"""
    action = request.form.get('action', '')

    if action == 'cancel':
        session.pop('teacher_import_data', None)
        flash('已取消导入', 'info')
        return redirect(url_for('admin.teacher_student_import'))

    if action != 'confirm':
        flash('无效的操作', 'error')
        return redirect(url_for('admin.teacher_student_import'))

    import_data = session.get('teacher_import_data')
    if not import_data:
        flash('导入数据已过期或不存在，请重新上传文件', 'error')
        return redirect(url_for('admin.teacher_student_import'))

    valid_users = import_data.get('valid_users', [])

    if not valid_users:
        flash('没有可导入的有效数据', 'error')
        return redirect(url_for('admin.teacher_student_import'))

    try:
        success_count = 0
        skip_count = 0
        error_count = 0

        for user_data in valid_users:
            try:
                if User.query.filter_by(username=user_data['username']).first():
                    skip_count += 1
                    continue

                email = user_data.get('email', '')
                if email and User.query.filter_by(email=email).first():
                    skip_count += 1
                    continue

                user = User(
                    username=user_data['username'],
                    email=email or f"{user_data['username']}@studyclash.local",
                    real_name=user_data['real_name'],
                    nickname=user_data.get('nickname', user_data['real_name']),
                    student_id=user_data.get('student_id') or None,
                    phone=user_data.get('phone') or None,
                    school=user_data.get('school') or None,
                    grade=user_data.get('grade') or None,
                    major=user_data.get('major') or None,
                    class_name=user_data.get('class_name') or None,
                    role='student',
                    is_active=True,
                    can_edit_profile=True,
                    created_by=current_user.id
                )

                user.set_password(user_data['password'])
                db.session.add(user)
                success_count += 1

            except Exception:
                error_count += 1
                continue

        db.session.commit()
        session.pop('teacher_import_data', None)

        log_operation('import', 'student',
                      detail=f'教师{current_user.username}导入{success_count}个学生(跳过{skip_count}失败{error_count})')

        message = f'导入完成！成功: {success_count} 个学生'
        if skip_count > 0:
            message += f' | 跳过重复: {skip_count} 个'
        if error_count > 0:
            message += f' | 错误: {error_count} 个'

        flash(message, 'success')

        breadcrumb = [
            {'label': '学生管理'},
            {'label': '学生导入'},
            {'label': '导入结果'}
        ]

        return render_template(
            'admin/teacher_student_import.html',
            breadcrumb=breadcrumb,
            step='result',
            success_count=success_count,
            skip_count=skip_count,
            error_count=error_count
        )

    except Exception as e:
        db.session.rollback()
        session.pop('teacher_import_data', None)
        flash(f'导入失败: {str(e)}', 'error')
        return redirect(url_for('admin.teacher_student_import'))


@admin_bp.route('/teacher/student-import/template')
@login_required
@role_required('teacher')
@teacher_permission_required('can_import_students')
def teacher_student_import_template():
    """下载导入模板"""
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "导入数据"

    headers = [
        ("用户名", "username"),
        ("真实姓名", "real_name"),
        ("密码", "password"),
        ("学号", "student_id"),
        ("邮箱", "email"),
        ("昵称", "nickname"),
        ("手机", "phone"),
        ("学校", "school"),
        ("年级", "grade"),
        ("专业", "major"),
        ("班级", "class_name")
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, (cn, en) in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=f"{cn}\n({en})")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    sample_data = [
        ["zhangsan", "张三", "password123", "2024001", "", "张三", "13800138000", "XX中学", "高一", "文科", "高一(1)班"],
        ["lisi", "李四", "password123", "2024002", "", "李四", "13800138001", "XX中学", "高二", "理科", "高二(3)班"]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws_data.column_dimensions[chr(64 + col)].width = 15

    ws_data.row_dimensions[1].height = 40

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='学生导入模板.xlsx'
    )

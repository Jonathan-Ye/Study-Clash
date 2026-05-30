import io
import re
from datetime import datetime
from flask import render_template, request, redirect, url_for, flash, send_file, session
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app import db
from app.models import User, DictionaryItem
from app.models.login_security import LoginAttempt
from app.routes.admin import admin_bp, admin_required
from app.utils.security import validate_password
from app.utils.op_log import log_operation

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
MAX_IMPORT_ROWS = 500

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@admin_bp.route('/users')
@login_required
@admin_required
def users():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in (10, 20, 50, 100, 200):
        per_page = 20
    search = request.args.get('search', '')
    school = request.args.get('school', '')
    grade = request.args.get('grade', '')
    major = request.args.get('major', '')
    class_name = request.args.get('class_name', '')
    role = request.args.get('role', '')
    can_edit = request.args.get('can_edit', '')
    is_active = request.args.get('is_active', '')

    query = User.query
    if search:
        query = query.filter(
            db.or_(
                User.username.contains(search),
                User.nickname.contains(search),
                User.email.contains(search),
                User.real_name.contains(search),
                User.student_id.contains(search)
            )
        )
    if school:
        query = query.filter_by(school=school)
    if grade:
        query = query.filter_by(grade=grade)
    if major:
        query = query.filter_by(major=major)
    if class_name:
        query = query.filter_by(class_name=class_name)
    if role:
        query = query.filter_by(role=role)
    if can_edit == '1':
        query = query.filter_by(can_edit_profile=True)
    elif can_edit == '0':
        query = query.filter_by(can_edit_profile=False)
    if is_active == '1':
        query = query.filter_by(is_active=True)
    elif is_active == '0':
        query = query.filter_by(is_active=False)

    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=per_page)

    # 加载字典选项
    dict_options = {
        'schools': DictionaryItem.get_options('school'),
        'grades': DictionaryItem.get_options('grade'),
        'majors': DictionaryItem.get_options('major'),
        'classes': DictionaryItem.get_options('class_name')
    }

    # 获取用户锁定状态
    user_lock_status = {}
    for user in pagination.items:
        if user.role != 'admin':
            is_locked, remaining = LoginAttempt.is_locked(user.username)
            if is_locked:
                user_lock_status[user.id] = {
                    'is_locked': True,
                    'remaining_minutes': remaining,
                    'fail_count': LoginAttempt.query.filter_by(username=user.username).first().fail_count if LoginAttempt.query.filter_by(username=user.username).first() else 0
                }
            else:
                attempt = LoginAttempt.query.filter_by(username=user.username).first()
                if attempt and attempt.fail_count > 0:
                    user_lock_status[user.id] = {
                        'is_locked': False,
                        'remaining_minutes': 0,
                        'fail_count': attempt.fail_count
                    }

    return render_template(
        'admin/users.html',
        pagination=pagination,
        search=search,
        per_page=per_page,
        school=school,
        grade=grade,
        major=major,
        class_name=class_name,
        dict_options=dict_options,
        user_lock_status=user_lock_status
    )

@admin_bp.route('/users/export')
@login_required
@admin_required
def export_users():
    """导出用户列表为Excel"""
    search = request.args.get('search', '')
    school = request.args.get('school', '')
    grade = request.args.get('grade', '')
    major = request.args.get('major', '')
    class_name = request.args.get('class_name', '')
    role = request.args.get('role', '')
    can_edit = request.args.get('can_edit', '')
    is_active = request.args.get('is_active', '')

    query = User.query
    if search:
        query = query.filter(
            db.or_(
                User.username.contains(search),
                User.nickname.contains(search),
                User.email.contains(search),
                User.real_name.contains(search),
                User.student_id.contains(search)
            )
        )
    if school:
        query = query.filter_by(school=school)
    if grade:
        query = query.filter_by(grade=grade)
    if major:
        query = query.filter_by(major=major)
    if class_name:
        query = query.filter_by(class_name=class_name)
    if role:
        query = query.filter_by(role=role)
    if can_edit == '1':
        query = query.filter_by(can_edit_profile=True)
    elif can_edit == '0':
        query = query.filter_by(can_edit_profile=False)
    if is_active == '1':
        query = query.filter_by(is_active=True)
    elif is_active == '0':
        query = query.filter_by(is_active=False)

    users = query.order_by(User.created_at.desc()).limit(10000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "用户列表"

    headers = ['用户名', '昵称', '真实姓名', '学号', '邮箱', '手机', '学校', '年级', '专业', '班级', '角色', '状态', '积分', '注册时间']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    role_map = {'admin': '管理员', 'teacher': '教师', 'student': '学生'}
    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=user.username)
        ws.cell(row=row_idx, column=2, value=user.nickname or '')
        ws.cell(row=row_idx, column=3, value=user.real_name or '')
        ws.cell(row=row_idx, column=4, value=user.student_id or '')
        ws.cell(row=row_idx, column=5, value=user.email or '')
        ws.cell(row=row_idx, column=6, value=user.phone or '')
        ws.cell(row=row_idx, column=7, value=user.school or '')
        ws.cell(row=row_idx, column=8, value=user.grade or '')
        ws.cell(row=row_idx, column=9, value=user.major or '')
        ws.cell(row=row_idx, column=10, value=user.class_name or '')
        ws.cell(row=row_idx, column=11, value=role_map.get(user.role, user.role))
        ws.cell(row=row_idx, column=12, value='正常' if user.is_active else '禁用')
        ws.cell(row=row_idx, column=13, value=user.total_points or 0)
        ws.cell(row=row_idx, column=14, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_operation('export', 'user', detail=f'导出 {len(users)} 条用户记录')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='用户列表.xlsx'
    )


@admin_bp.route('/users/<int:user_id>/toggle-admin', methods=['POST'])
@login_required
@admin_required
def toggle_admin(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('不能修改自己的管理员权限', 'error')
        return redirect(url_for('admin.users'))

    user.role = 'admin' if user.role != 'admin' else 'student'
    db.session.commit()

    action = '设置为管理员' if user.role == 'admin' else '取消管理员权限'
    log_operation('update', 'user', target_id=user.id, target_name=user.nickname or user.username,
                 detail=f'{action}')
    flash(f'已{action}: {user.nickname or user.username}', 'success')
    return redirect(url_for('admin.users'))

@admin_bp.route('/users/<int:user_id>/toggle-active', methods=['POST'])
@login_required
@admin_required
def toggle_active(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('不能禁用自己的账号', 'error')
        return redirect(url_for('admin.users'))

    user.is_active = not user.is_active
    db.session.commit()

    action = '启用' if user.is_active else '禁用'
    log_operation('update', 'user', target_id=user.id, target_name=user.nickname or user.username,
                 detail=f'{action}账号')
    flash(f'已{action}账号: {user.nickname or user.username}', 'success')
    return redirect(url_for('admin.users'))

@admin_bp.route('/users/<int:user_id>/unlock', methods=['POST'])
@login_required
@admin_required
def unlock_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.role == 'admin':
        flash('管理员账户不受锁定限制', 'info')
        return redirect(url_for('admin.users'))

    LoginAttempt.reset(user.username)
    log_operation('unlock', 'user', target_id=user.id, target_name=user.nickname or user.username,
                 detail='管理员手动解锁账户')
    flash(f'已解锁用户: {user.nickname or user.username}', 'success')
    return redirect(url_for('admin.users'))

@admin_bp.route('/users/create', methods=['POST'])
@login_required
@admin_required
def create_user():
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        real_name = request.form.get('real_name', '').strip()
        email = request.form.get('email', '').strip()
        nickname = request.form.get('nickname', '').strip()
        student_id = request.form.get('student_id', '').strip()
        phone = request.form.get('phone', '').strip()
        school = request.form.get('school', '').strip()
        grade = request.form.get('grade', '').strip()
        major = request.form.get('major', '').strip()
        class_name = request.form.get('class_name', '').strip()

        role = request.form.get('role', 'student')
        # 确保role值合法
        if role not in ('admin', 'teacher', 'student'):
            role = 'student'
        from flask import current_app
        current_app.logger.info(f'[create_user] form role={request.form.get("role")}, final role={role}, all form keys={list(request.form.keys())}')
        can_edit_profile = request.form.get('can_edit_profile') == 'on'
        is_active = request.form.get('is_active') == 'on'
        
        # 教师细粒度权限
        can_manage_subjects = request.form.get('can_manage_subjects') == 'on'
        can_manage_chapters = request.form.get('can_manage_chapters') == 'on'
        can_manage_questions = request.form.get('can_manage_questions') == 'on'
        can_import_questions = request.form.get('can_import_questions') == 'on'
        can_export_questions = request.form.get('can_export_questions') == 'on'
        can_import_students = request.form.get('can_import_students') == 'on'
        can_view_student_analysis = request.form.get('can_view_student_analysis') == 'on'
        can_view_knowledge_analysis = request.form.get('can_view_knowledge_analysis') == 'on'

        # 后端验证 - 必填项检查
        errors = []

        # 用户名验证
        if not username:
            errors.append('用户名为必填项')
        elif len(username) < 3:
            errors.append('用户名至少需要3个字符')
        elif len(username) > 64:
            errors.append('用户名不能超过64个字符')
        elif not re.match(r'^[a-zA-Z0-9_]+$', username):
            errors.append('用户名只能包含字母、数字和下划线')

        # 检查用户名是否已存在
        if username and User.query.filter_by(username=username).first():
            errors.append(f'用户名 "{username}" 已被占用')

        # 密码验证
        if not password:
            errors.append('密码为必填项')
        else:
            is_valid, pwd_errors = validate_password(password)
            if not is_valid:
                errors.extend(pwd_errors)
            elif len(password) > 128:
                errors.append('密码不能超过128个字符')

        # 姓名验证
        if not real_name:
            errors.append('真实姓名为必填项')
        elif len(real_name) > 64:
            errors.append('姓名不能超过64个字符')

        # 学号验证（必填）
        if not student_id:
            errors.append('学号为必填项')
        elif len(student_id) > 32:
            errors.append('学号不能超过32位数字')
        elif not re.match(r'^[0-9]+$', student_id):
            errors.append('学号必须是纯数字')

        # 检查学号是否已存在
        if student_id and User.query.filter_by(student_id=student_id).first():
            errors.append(f'学号 "{student_id}" 已被使用')

        # 邮箱验证（如果填写了）
        if email:
            if len(email) > 120:
                errors.append('邮箱不能超过120个字符')
            elif not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
                errors.append('邮箱格式不正确')
            elif User.query.filter_by(email=email).first():
                errors.append(f'邮箱 "{email}" 已被使用')

        # 学号验证（如果填写了）
        if student_id:
            if not re.match(r'^[0-9]+$', student_id):
                errors.append('学号必须是纯数字')
            elif len(student_id) > 32:
                errors.append('学号不能超过32位数字')

        # 手机号验证（如果填写了）
        if phone and not re.match(r'^[0-9+\- ]*$', phone):
            errors.append('手机号码格式不正确')

        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('admin.users'))

        # 处理可选字段默认值
        if not email:
            email = f"{username}@studyclash.local"

        if not nickname:
            nickname = real_name

        # 创建用户对象
        user = User(
            username=username,
            email=email,
            real_name=real_name,
            nickname=nickname,
            student_id=student_id or None,
            phone=phone or None,
            school=school or None,
            grade=grade or None,
            major=major or None,
            class_name=class_name or None,
            role=role,
            is_active=is_active,
            can_edit_profile=can_edit_profile,
            can_manage_subjects=can_manage_subjects,
            can_manage_chapters=can_manage_chapters,
            can_manage_questions=can_manage_questions,
            can_import_questions=can_import_questions,
            can_export_questions=can_export_questions,
            can_import_students=can_import_students,
            can_view_student_analysis=can_view_student_analysis,
            can_view_knowledge_analysis=can_view_knowledge_analysis
        )

        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        log_operation('create', 'user', target_id=user.id, target_name=real_name,
                     detail=f'username={username}, role={role}')
        flash(f'✅ 成功创建用户: {real_name} ({username})', 'success')
        return redirect(url_for('admin.users'))

    except Exception as e:
        db.session.rollback()
        flash(f'❌ 创建用户失败: {str(e)}', 'error')
        return redirect(url_for('admin.users'))

@admin_bp.route('/users/batch', methods=['POST'])
@login_required
@admin_required
def batch_users():
    action = request.form.get('action')
    user_ids = request.form.get('user_ids', '[]')

    try:
        import json
        user_ids = json.loads(user_ids)
    except:
        flash('无效的用户ID', 'error')
        return redirect(url_for('admin.users'))

    if not user_ids:
        flash('请选择要操作的用户', 'error')
        return redirect(url_for('admin.users'))

    users = User.query.filter(User.id.in_(user_ids)).all()
    if not users:
        flash('未找到选中的用户', 'error')
        return redirect(url_for('admin.users'))

    count = len(users)
    success_count = 0

    if action == 'activate':
        for user in users:
            if user.id != current_user.id:
                user.is_active = True
                success_count += 1
        db.session.commit()
        log_operation('batch_operation', 'user', detail=f'激活 {success_count} 个用户')
        flash(f'已成功激活 {success_count} 个用户', 'success')

    elif action == 'deactivate':
        for user in users:
            if user.id != current_user.id:
                user.is_active = False
                success_count += 1
        db.session.commit()
        log_operation('batch_operation', 'user', detail=f'禁用 {success_count} 个用户')
        flash(f'已成功禁用 {success_count} 个用户', 'success')

    elif action == 'lock_profile':
        for user in users:
            if user.id != current_user.id:
                user.can_edit_profile = False
                success_count += 1
        db.session.commit()
        flash(f'已禁止 {success_count} 个用户修改资料', 'success')

    elif action == 'unlock_profile':
        for user in users:
            user.can_edit_profile = True
            success_count += 1
        db.session.commit()
        flash(f'已允许 {success_count} 个用户修改资料', 'success')

    elif action == 'set_admin':
        for user in users:
            if user.id != current_user.id and user.role != 'admin':
                user.role = 'admin'
                success_count += 1
        db.session.commit()
        flash(f'已将 {success_count} 个用户设为管理员', 'success')

    elif action == 'remove_admin':
        for user in users:
            if user.id != current_user.id and user.role == 'admin':
                user.role = 'student'
                success_count += 1
        db.session.commit()
        flash(f'已取消 {success_count} 个用户的管理员权限', 'success')

    elif action == 'set_teacher':
        for user in users:
            if user.id != current_user.id and user.role != 'admin':
                user.role = 'teacher'
                success_count += 1
        db.session.commit()
        flash(f'已将 {success_count} 个用户设为教师', 'success')

    elif action == 'remove_teacher':
        for user in users:
            if user.id != current_user.id and user.role == 'teacher':
                user.role = 'student'
                success_count += 1
        db.session.commit()
        flash(f'已取消 {success_count} 个用户的教师权限', 'success')

    elif action == 'delete':
        from app.models.points import DailyStats, PointRecord, Leaderboard
        from app.models.game import GameRecord, GamePlayer
        from app.models.wrong_question import WrongQuestion, ReviewStreak, WrongQuestionCollection, WrongQuestionCollectionItem, WrongQuestionNote, ChallengeProgress
        from app.models.question import Question, Subject, UserAnswer
        from app.models.ranks import TierPromotionHistory
        from app.models.achievements import RankHistory

        for user in users:
            if user.id != current_user.id:
                uid = user.id
                # 删除关联数据
                DailyStats.query.filter_by(user_id=uid).delete()
                PointRecord.query.filter_by(user_id=uid).delete()
                GameRecord.query.filter_by(user_id=uid).delete()
                GamePlayer.query.filter_by(user_id=uid).delete()
                WrongQuestionNote.query.filter_by(user_id=uid).delete()
                db.session.execute(db.text(
                    'DELETE FROM wrong_question_collection_items WHERE wrong_question_id IN '
                    '(SELECT id FROM wrong_questions WHERE user_id = :uid)'
                ), {'uid': uid})
                WrongQuestion.query.filter_by(user_id=uid).delete()
                WrongQuestionCollection.query.filter_by(user_id=uid).delete()
                ReviewStreak.query.filter_by(user_id=uid).delete()
                ChallengeProgress.query.filter_by(user_id=uid).delete()
                UserAnswer.query.filter_by(user_id=uid).delete()
                TierPromotionHistory.query.filter_by(user_id=uid).delete()
                RankHistory.query.filter_by(user_id=uid).delete()
                Leaderboard.query.filter_by(user_id=uid).delete()
                Question.query.filter_by(created_by=uid).update({'created_by': None})
                Subject.query.filter_by(created_by=uid).update({'created_by': None})
                User.query.filter_by(created_by=uid).update({'created_by': None})
                db.session.execute(db.text('UPDATE admin_logs SET admin_id = NULL WHERE admin_id = :uid'), {'uid': uid})
                db.session.execute(db.text('UPDATE announcements SET created_by = NULL WHERE created_by = :uid'), {'uid': uid})
                db.session.execute(db.text('UPDATE game_rooms SET created_by = NULL WHERE created_by = :uid'), {'uid': uid})
                db.session.delete(user)
                success_count += 1
        db.session.commit()
        flash(f'已成功删除 {success_count} 个用户', 'success')

    else:
        flash('未知的操作类型', 'error')

    return redirect(url_for('admin.users'))

@admin_bp.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    user = User.query.get_or_404(id)

    # 处理POST请求（保存修改）
    if request.method == 'POST':
        try:
            # 获取表单数据
            user.username = request.form.get('username', '').strip()
            user.email = request.form.get('email', '').strip()
            user.nickname = request.form.get('nickname', '').strip()
            user.real_name = request.form.get('real_name', '').strip()
            user.student_id = request.form.get('student_id', '').strip() or None
            user.phone = request.form.get('phone', '').strip() or None
            user.school = request.form.get('school', '').strip() or None
            user.grade = request.form.get('grade', '').strip() or None
            user.major = request.form.get('major', '').strip() or None
            user.class_name = request.form.get('class_name', '').strip() or None
            
            # 密码（如果填写了）
            new_password = request.form.get('new_password', '').strip()
            if new_password:
                is_valid, pwd_errors = validate_password(new_password)
                if not is_valid:
                    for err in pwd_errors:
                        flash(err, 'error')
                else:
                    user.set_password(new_password)
            
            # 账户设置
            role = request.form.get('role', 'student')
            # 确保role值合法
            if role not in ('admin', 'teacher', 'student'):
                role = 'student'
            from flask import current_app
            current_app.logger.info(f'[edit_user] form role={request.form.get("role")}, final role={role}, current role={user.role}, all form keys={list(request.form.keys())}')
            # 不允许降级管理员角色（除非是自己）
            if user.role == 'admin' and role != 'admin' and user.id == current_user.id:
                flash('不能修改自己的管理员角色', 'error')
            else:
                user.role = role
            user.is_active = request.form.get('is_active') == 'on'
            user.can_edit_profile = request.form.get('can_edit_profile') == 'on'
            user.participate_in_games = request.form.get('participate_in_games') == 'on'
            user.show_in_leaderboard = request.form.get('show_in_leaderboard') == 'on'
            
            # 教师细粒度权限
            user.can_manage_subjects = request.form.get('can_manage_subjects') == 'on'
            user.can_manage_chapters = request.form.get('can_manage_chapters') == 'on'
            user.can_manage_questions = request.form.get('can_manage_questions') == 'on'
            user.can_import_questions = request.form.get('can_import_questions') == 'on'
            user.can_export_questions = request.form.get('can_export_questions') == 'on'
            user.can_import_students = request.form.get('can_import_students') == 'on'
            user.can_view_student_analysis = request.form.get('can_view_student_analysis') == 'on'
            user.can_view_knowledge_analysis = request.form.get('can_view_knowledge_analysis') == 'on'
            
            # 验证必填字段
            errors = []
            
            if not user.username:
                errors.append('用户名为必填项')
            elif len(user.username) < 3 or len(user.username) > 64:
                errors.append('用户名长度必须在3-64字符之间')
            elif not re.match(r'^[a-zA-Z0-9_]+$', user.username):
                errors.append('用户名只能包含字母、数字和下划线')
                
            if not user.real_name:
                errors.append('真实姓名为必填项')
                
            if not user.email:
                errors.append('邮箱为必填项')
            elif not re.match(r'^[^@]+@[^@]+\.[^@]+$', user.email):
                errors.append('邮箱格式不正确')
                
            if user.student_id and not re.match(r'^[0-9]+$', user.student_id):
                errors.append('学号必须是纯数字')
            
            if errors:
                for error in errors:
                    flash(error, 'error')
            else:
                db.session.commit()
                flash(f'✅ 用户 "{user.nickname or user.username}" 的信息已更新！', 'success')
                return redirect(url_for('admin.users'))
                
        except Exception as e:
            db.session.rollback()
            flash(f'❌ 更新失败: {str(e)}', 'error')

    # 加载字典选项
    dict_options = {
        'schools': DictionaryItem.get_options('school'),
        'grades': DictionaryItem.get_options('grade'),
        'majors': DictionaryItem.get_options('major'),
        'classes': DictionaryItem.get_options('class_name')
    }

    return render_template(
        'admin/edit_user.html',
        user=user,
        dict_options=dict_options
    )

@admin_bp.route('/users/<int:id>/adjust-points', methods=['POST'])
@login_required
@admin_required
def adjust_user_points(id):
    try:
        points = int(request.form.get('points', 0))
        reason = request.form.get('reason', 'admin_add')

        user = User.query.get_or_404(id)
        
        if reason == 'correction':
            user.set_points(points, reason)
            flash(f'已将用户 {user.nickname or user.username} 的积分修正为 {points}', 'success')
        else:
            user.add_points(points, reason)
            action = '增加' if points > 0 else '扣除'
            flash(f'已{action} {abs(points)} 积分给用户: {user.nickname or user.username}', 'success')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'调整积分失败: {str(e)}', 'error')

    return redirect(url_for('admin.edit_user', id=id))

@admin_bp.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('不能删除自己', 'error')
        return redirect(url_for('admin.users'))

    try:
        # 先删除所有关联数据，避免外键约束错误
        from app.models.points import DailyStats, PointRecord, Leaderboard
        from app.models.game import GameRecord, GamePlayer
        from app.models.wrong_question import WrongQuestion, ReviewStreak, WrongQuestionCollection, WrongQuestionCollectionItem, WrongQuestionNote, ChallengeProgress
        from app.models.question import Question, Subject, UserAnswer
        from app.models.ranks import TierPromotionHistory
        from app.models.achievements import RankHistory

        # 删除用户相关的所有数据（按依赖顺序）
        DailyStats.query.filter_by(user_id=id).delete()
        PointRecord.query.filter_by(user_id=id).delete()
        GameRecord.query.filter_by(user_id=id).delete()
        GamePlayer.query.filter_by(user_id=id).delete()
        WrongQuestionNote.query.filter_by(user_id=id).delete()
        db.session.execute(db.text(
            'DELETE FROM wrong_question_collection_items WHERE wrong_question_id IN '
            '(SELECT id FROM wrong_questions WHERE user_id = :uid)'
        ), {'uid': id})
        WrongQuestion.query.filter_by(user_id=id).delete()
        WrongQuestionCollection.query.filter_by(user_id=id).delete()
        ReviewStreak.query.filter_by(user_id=id).delete()
        ChallengeProgress.query.filter_by(user_id=id).delete()
        UserAnswer.query.filter_by(user_id=id).delete()
        TierPromotionHistory.query.filter_by(user_id=id).delete()
        RankHistory.query.filter_by(user_id=id).delete()
        Leaderboard.query.filter_by(user_id=id).delete()

        # 将nullable外键设为NULL（不删除这些记录）
        Question.query.filter_by(created_by=id).update({'created_by': None})
        Subject.query.filter_by(created_by=id).update({'created_by': None})
        User.query.filter_by(created_by=id).update({'created_by': None})
        # admin_logs和announcements的created_by/admin_id也设为NULL
        db.session.execute(db.text('UPDATE admin_logs SET admin_id = NULL WHERE admin_id = :uid'), {'uid': id})
        db.session.execute(db.text('UPDATE announcements SET created_by = NULL WHERE created_by = :uid'), {'uid': id})
        db.session.execute(db.text('UPDATE game_rooms SET created_by = NULL WHERE created_by = :uid'), {'uid': id})

        db.session.delete(user)
        db.session.commit()
        log_operation('delete', 'user', target_id=id, target_name=user.nickname or user.username)
        flash(f'已删除用户: {user.nickname or user.username}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除用户失败: {str(e)}', 'error')

    return redirect(url_for('admin.users'))

@admin_bp.route('/users/import', methods=['POST'])
@login_required
@admin_required
def import_users():
    if 'file' not in request.files:
        flash('未选择文件', 'error')
        return redirect(url_for('admin.users'))

    file = request.files['file']
    if file.filename == '':
        flash('未选择文件', 'error')
        return redirect(url_for('admin.users'))

    if not allowed_file(file.filename):
        flash('不支持的文件格式，请上传 .xlsx, .xls 或 .csv 文件', 'error')
        return redirect(url_for('admin.users'))

    try:
        # 获取表单选项
        is_active = request.form.get('is_active') == 'on'
        can_edit_profile = request.form.get('can_edit_profile') == 'on'
        set_admin = request.form.get('set_admin') == 'on'

        # 读取文件内容
        if file.filename.endswith('.csv'):
            import pandas as pd
            df = pd.read_csv(file, encoding='utf-8-sig')
        else:
            import pandas as pd
            df = pd.read_excel(file)

        def _safe_str(val):
            if pd.isna(val):
                return ''
            return str(val).strip()

        def _safe_int(val, default=0):
            if pd.isna(val):
                return default
            return int(val)

        # 数据量限制
        if len(df) > MAX_IMPORT_ROWS:
            flash(f'数据量过大！单次最多导入 {MAX_IMPORT_ROWS} 条记录，当前有 {len(df)} 条', 'error')
            return redirect(url_for('admin.users'))

        # 标准化列名（处理可能的中文/英文列名）
        column_mapping = {
            'username': ['username', '用户名', '账号', '登录名'],
            'real_name': ['real_name', '姓名', '真实姓名', 'name'],
            'password': ['password', '密码', '登陆密码'],
            'email': ['email', '邮箱', '电子邮件', 'Email'],
            'nickname': ['nickname', '昵称', '别名'],
            'student_id': ['student_id', '学号'],
            'phone': ['phone', '手机', '手机号码', '电话'],
            'school': ['school', '学校'],
            'grade': ['grade', '年级'],
            'major': ['major', '专业', '文理科'],
            'class_name': ['class_name', '班级', 'ClassName']
        }

        # 反转映射用于查找标准列名
        reverse_mapping = {}
        for standard_col, variants in column_mapping.items():
            for variant in variants:
                reverse_mapping[variant.lower()] = standard_col

        # 重命名列（更宽松的匹配：去除空格、下划线、换行符）
        new_columns = {}
        for col in df.columns:
            col_str = str(col).strip().lower()
            col_parts = [p.strip() for p in col_str.replace('\n', ' ').replace('_', ' ').split(' ') if p.strip()]
            matched = False
            for variant_lower, standard_col in reverse_mapping.items():
                variant_clean = variant_lower.replace(' ', '').replace('_', '')
                col_clean = col_str.replace(' ', '').replace('_', '').replace('\n', '')
                if col_clean == variant_clean:
                    new_columns[col] = standard_col
                    matched = True
                    break
                for part in col_parts:
                    part_clean = part.replace(' ', '').replace('_', '')
                    if part_clean == variant_clean:
                        new_columns[col] = standard_col
                        matched = True
                        break
                if matched:
                    break
            if not matched:
                new_columns[col] = col_str.replace('\n', '_').replace(' ', '_')

        df.rename(columns=new_columns, inplace=True)

        # 预加载字典选项用于验证和确认页面
        dict_options = {
            'school_values': {item.value: item.label for item in DictionaryItem.get_options('school')},
            'grade_values': {item.value: item.label for item in DictionaryItem.get_options('grade')},
            'major_values': {item.value: item.label for item in DictionaryItem.get_options('major')},
            'class_values': {item.value: item.label for item in DictionaryItem.get_options('class_name')},
            'schools': DictionaryItem.get_options('school'),
            'grades': DictionaryItem.get_options('grade'),
            'majors': DictionaryItem.get_options('major'),
            'classes': DictionaryItem.get_options('class_name')
        }

        valid_users = []
        invalid_users = []
        total_rows = len(df)

        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel行号从2开始（第1行是标题）

            user_data = {
                'row': row_num,
                'username': _safe_str(row.get('username')),
                'real_name': _safe_str(row.get('real_name')),
                'password': _safe_str(row.get('password')),
                'email': _safe_str(row.get('email')),
                'nickname': _safe_str(row.get('nickname')),
                'student_id': _safe_str(row.get('student_id')),
                'phone': _safe_str(row.get('phone')),
                'school': _safe_str(row.get('school')),
                'grade': _safe_str(row.get('grade')),
                'major': _safe_str(row.get('major')),
                'class_name': _safe_str(row.get('class_name'))
            }

            errors = []

            # 必填项验证
            if not user_data['username']:
                errors.append('用户名为空')
            elif len(user_data['username']) < 3:
                errors.append('用户名太短（至少3个字符）')
            elif not re.match(r'^[a-zA-Z0-9_]+$', user_data['username']):
                errors.append('用户名格式不正确')

            if not user_data['real_name']:
                errors.append('姓名为空')

            if not user_data['password']:
                errors.append('密码为空')
            else:
                is_valid, pwd_errors = validate_password(user_data['password'])
                if not is_valid:
                    errors.extend(pwd_errors)

            # 学号验证（必填）
            if not user_data['student_id']:
                errors.append('学号为空')
            elif not re.match(r'^[0-9]+$', user_data['student_id']):
                errors.append('学号必须是纯数字')

            # 唯一性检查
            if user_data['username'] and User.query.filter_by(username=user_data['username']).first():
                errors.append(f'用户名 "{user_data["username"]}" 已存在')

            # 邮箱验证
            if user_data['email']:
                if not re.match(r'^[^@]+@[^@]+\.[^@]+$', user_data['email']):
                    errors.append('邮箱格式不正确')
                elif User.query.filter_by(email=user_data['email']).first():
                    errors.append(f'邮箱 "{user_data["email"]}" 已存在')

            # 学号验证
            if user_data['student_id'] and not re.match(r'^[0-9]+$', user_data['student_id']):
                errors.append('学号必须是纯数字')

            if errors:
                invalid_users.append({
                    'row': row_num,
                    'real_name': user_data['real_name'] or '-',
                    'student_id': user_data['student_id'] or '-',
                    'errors': errors
                })
            else:
                # 处理可选字段默认值
                if not user_data['email']:
                    user_data['email'] = f"{user_data['username']}@studyclash.local"
                if not user_data['nickname']:
                    user_data['nickname'] = user_data['real_name']

                # 字典字段校验（警告而非错误，允许导入但提示用户）
                warnings = []
                dict_fields = [
                    ('school', dict_options['school_values'], '学校'),
                    ('grade', dict_options['grade_values'], '年级'),
                    ('major', dict_options['major_values'], '专业'),
                    ('class_name', dict_options['class_values'], '班级')
                ]
                for field_key, valid_values, field_label in dict_fields:
                    value = user_data.get(field_key, '')
                    if value and value not in valid_values:
                        available = ', '.join(sorted(valid_values.keys())) if valid_values else '（字典中暂无数据，请先在字典管理中配置）'
                        warnings.append({
                            'field': field_key,
                            'label': field_label,
                            'message': f'{field_label} "{value}" 不在字典中，请在确认页面修改',
                            'available': available
                        })

                user_data['warnings'] = warnings
                valid_users.append(user_data)

        # 存储数据到session供确认页面使用
        session['import_data'] = {
            'valid_users': valid_users,
            'invalid_users': invalid_users,
            'total': total_rows,
            'settings': {
                'is_active': is_active,
                'can_edit_profile': can_edit_profile,
                'set_admin': set_admin
            },
            'dict_options': {
                'schools': [{'value': item.value, 'label': item.label} for item in dict_options['schools']],
                'grades': [{'value': item.value, 'label': item.label} for item in dict_options['grades']],
                'majors': [{'value': item.value, 'label': item.label} for item in dict_options['majors']],
                'classes': [{'value': item.value, 'label': item.label} for item in dict_options['classes']]
            }
        }

        return render_template(
            'admin/import_confirm.html',
            total=total_rows,
            valid_users=valid_users,
            invalid_users=invalid_users,
            dict_options=dict_options
        )

    except Exception as e:
        flash(f'解析文件失败: {str(e)}', 'error')
        return redirect(url_for('admin.users'))

@admin_bp.route('/users/import-template')
@login_required
@admin_required
def download_import_template():
    wb = Workbook()

    ws_data = wb.active
    ws_data.title = "导入数据"

    headers = [
        ("用户名", "username"),
        ("真实姓名", "real_name"),
        ("密码", "password"),
        ("学号", "student_id"),
        ("邮箱", "email"),
        ("昵称", "nickname"),
        ("手机", "phone"),
        ("学校", "school"),
        ("年级", "grade"),
        ("专业", "major"),
        ("班级", "class_name")
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, (cn, en) in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=f"{cn}\n({en})")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    sample_data = [
        ["zhangsan", "张三", "password123", "2024001", "", "张三", "13800138000", "XX中学", "高一", "文科", "高一(1)班"],
        ["lisi", "李四", "password123", "2024002", "", "李四", "13800138001", "XX中学", "高二", "理科", "高二(3)班"]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(col)].width = 15

    ws_data.row_dimensions[1].height = 40

    ws_guide = wb.create_sheet(title="填写说明")

    guide_content = [
        ["【必填字段】（4项）", ""],
        ["用户名(username)", "用于登录系统，必须唯一（英文/数字/下划线，3-64字符）"],
        ["真实姓名(real_name)", "学生的真实姓名（支持中文）"],
        ["密码(password)", "登录密码（建议6位以上）"],
        ["学号(student_id)", "学生唯一标识（必须是纯数字，不能重复）"],
        ["", ""],
        ["【可选字段】（7项）", ""],
        ["邮箱(email)", "不填则自动生成：用户名@studyclash.local"],
        ["昵称(nickname)", "不填则使用真实姓名"],
        ["手机(phone)", "手机号码（选填）"],
        ["学校(school)", "需与字典管理中的选项一致（选填）"],
        ["年级(grade)", "如：高一、高二、高三（选填）"],
        ["专业(major)", "如：文科、理科（选填）"],
        ["班级(class_name)", "如：高一(1)班（选填）"],
        ["", ""],
        ["【注意事项】", ""],
        ["1.", "每行代表一个用户，第一行为表头（不要删除）"],
        ["2.", "用户名、学号、邮箱不能重复，否则该行会被跳过"],
        ["3.", "密码至少6位字符"],
        ["4.", "学号必须是纯数字"],
        ["5.", "填写完成后保存为 .xlsx 或 .xls 格式即可导入"],
    ]

    title_font = Font(bold=True, size=12, color="4472C4")
    normal_font = Font(size=10)

    for row_idx, (field, desc) in enumerate(guide_content, 1):
        cell_field = ws_guide.cell(row=row_idx, column=1, value=field)
        cell_desc = ws_guide.cell(row=row_idx, column=2, value=desc)

        if field.startswith("【"):
            cell_field.font = title_font
        else:
            cell_field.font = normal_font
        cell_desc.font = normal_font

    ws_guide.column_dimensions['A'].width = 25
    ws_guide.column_dimensions['B'].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='用户批量导入模板.xlsx'
    )

@admin_bp.route('/users/import-confirm', methods=['POST'])
@login_required
@admin_required
def confirm_import():
    action = request.form.get('action', '')

    if action == 'cancel':
        session.pop('import_data', None)
        flash('已取消导入', 'info')
        return redirect(url_for('admin.users'))

    if action != 'confirm':
        flash('无效的操作', 'error')
        return redirect(url_for('admin.users'))

    import_data = session.get('import_data')
    if not import_data:
        flash('导入数据已过期或不存在，请重新上传文件', 'error')
        return redirect(url_for('admin.users'))

    valid_users = import_data.get('valid_users', [])
    settings = import_data.get('settings', {})

    if not valid_users:
        flash('没有可导入的有效数据', 'error')
        return redirect(url_for('admin.users'))

    try:
        success_count = 0
        skip_count = 0
        error_count = 0

        for user_data in valid_users:
            try:
                # 接受用户在确认页面编辑的字典字段
                row = user_data.get('row', 0)
                user_data['school'] = request.form.get(f'school_{row}', user_data.get('school') or '') or None
                user_data['grade'] = request.form.get(f'grade_{row}', user_data.get('grade') or '') or None
                user_data['major'] = request.form.get(f'major_{row}', user_data.get('major') or '') or None
                user_data['class_name'] = request.form.get(f'class_name_{row}', user_data.get('class_name') or '') or None

                # 再次检查唯一性（防止并发问题）
                if User.query.filter_by(username=user_data['username']).first():
                    skip_count += 1
                    continue

                email = user_data.get('email', '')
                if email and User.query.filter_by(email=email).first():
                    skip_count += 1
                    continue

                # 创建用户
                user = User(
                    username=user_data['username'],
                    email=email or f"{user_data['username']}@studyclash.local",
                    real_name=user_data['real_name'],
                    nickname=user_data.get('nickname', user_data['real_name']),
                    student_id=user_data.get('student_id') or None,
                    phone=user_data.get('phone') or None,
                    school=user_data.get('school') or None,
                    grade=user_data.get('grade') or None,
                    major=user_data.get('major') or None,
                    class_name=user_data.get('class_name') or None,
                    role='admin' if settings.get('set_admin', False) else 'student',
                    is_active=settings.get('is_active', True),
                    can_edit_profile=settings.get('can_edit_profile', True)
                )

                user.set_password(user_data['password'])
                db.session.add(user)
                success_count += 1

            except Exception as e:
                error_count += 1
                continue

        db.session.commit()
        session.pop('import_data', None)

        message = f'✅ 导入完成！成功: {success_count} 个用户'
        if skip_count > 0:
            message += f' | ⚠️ 跳过重复: {skip_count} 个'
        if error_count > 0:
            message += f' | ❌ 错误: {error_count} 个'

        flash(message, 'success')
        return redirect(url_for('admin.users'))

    except Exception as e:
        db.session.rollback()
        session.pop('import_data', None)
        flash(f'❌ 导入失败: {str(e)}', 'error')
        return redirect(url_for('admin.users'))

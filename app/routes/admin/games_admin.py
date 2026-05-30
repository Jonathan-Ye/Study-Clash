import io
from flask import render_template, request, jsonify, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app import db
from app.models import GameRoom, GamePlayer, GameRecord, User, PointRecord
from app.routes.admin import admin_bp, admin_required
from app.utils.op_log import log_operation

@admin_bp.route('/games')
@login_required
@admin_required
def games():
    page = request.args.get('page', 1, type=int)
    game_type = request.args.get('game_type')
    status = request.args.get('status')
    
    query = GameRoom.query
    
    if game_type:
        query = query.filter_by(game_type=game_type)
    if status:
        query = query.filter_by(status=status)
    
    pagination = query.order_by(GameRoom.created_at.desc()).paginate(page=page, per_page=20)
    
    return render_template('admin/games.html', pagination=pagination)


@admin_bp.route('/games/export')
@login_required
@admin_required
def export_games():
    """导出游戏记录为Excel"""
    game_type = request.args.get('game_type')
    status = request.args.get('status')

    query = GameRoom.query
    if game_type:
        query = query.filter_by(game_type=game_type)
    if status:
        query = query.filter_by(status=status)

    rooms = query.order_by(GameRoom.created_at.desc()).limit(10000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "游戏记录"

    headers = ['房间号', '游戏类型', '学科', '创建者', '玩家数', '状态', '创建时间', '结束时间']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    game_type_map = {'single': '单人挑战', 'battle': '双人对战', 'four': '四人挑战'}
    status_map = {'waiting': '等待中', 'playing': '进行中', 'finished': '已结束'}

    for row_idx, room in enumerate(rooms, 2):
        ws.cell(row=row_idx, column=1, value=room.room_code)
        ws.cell(row=row_idx, column=2, value=game_type_map.get(room.game_type, room.game_type))
        ws.cell(row=row_idx, column=3, value=room.subject.name if room.subject else '')
        creator = User.query.get(room.created_by) if room.created_by else None
        ws.cell(row=row_idx, column=4, value=creator.nickname or creator.username if creator else '')
        ws.cell(row=row_idx, column=5, value=room.current_players or 0)
        ws.cell(row=row_idx, column=6, value=status_map.get(room.status, room.status))
        ws.cell(row=row_idx, column=7, value=room.created_at.strftime('%Y-%m-%d %H:%M:%S') if room.created_at else '')
        ws.cell(row=row_idx, column=8, value=room.ended_at.strftime('%Y-%m-%d %H:%M:%S') if room.ended_at else '')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_operation('export', 'game', detail=f'导出 {len(rooms)} 条游戏记录')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='游戏记录.xlsx'
    )


@admin_bp.route('/games/room/<int:room_id>')
@login_required
@admin_required
def game_room_detail(room_id):
    room = GameRoom.query.get_or_404(room_id)
    players = GamePlayer.query.filter_by(room_id=room_id).all()
    records = GameRecord.query.filter_by(room_id=room_id).order_by(GameRecord.rank).all()
    
    player_details = []
    for p in players:
        user_records = [r for r in records if r.user_id == p.user_id]
        record = user_records[0] if user_records else None
        player_details.append({
            'player': p,
            'record': record,
            'user': p.user,
            'total_points_earned': sum(r.points_earned or 0 for r in user_records) if user_records else 0
        })
    
    return render_template('admin/game_room_detail.html', 
                          room=room, 
                          player_details=player_details)


@admin_bp.route('/games/user/<int:user_id>')
@login_required
@admin_required
def game_user_detail(user_id):
    page = request.args.get('page', 1, type=int)
    user = User.query.get_or_404(user_id)
    
    records = GameRecord.query.filter_by(user_id=user_id)\
        .order_by(GameRecord.created_at.desc()).paginate(page=page, per_page=20)
    
    total_games = GameRecord.query.filter_by(user_id=user_id).count()
    total_points = db.session.query(db.func.sum(GameRecord.points_earned))\
        .filter_by(user_id=user_id).scalar() or 0
    
    stats = {
        'total_games': total_games,
        'total_points': int(total_points),
        'single_count': GameRecord.query.filter_by(user_id=user_id, game_type='single').count(),
        'battle_count': GameRecord.query.filter_by(user_id=user_id, game_type='battle').count(),
        'four_count': GameRecord.query.filter_by(user_id=user_id, game_type='four').count(),
        'first_place': GameRecord.query.filter_by(user_id=user_id, rank=1).count(),
        'avg_score': db.session.query(db.func.avg(GameRecord.score))\
            .filter_by(user_id=user_id).scalar() or 0,
    }
    
    return render_template('admin/game_user_detail.html',
                          user=user,
                          records=records,
                          stats=stats)


@admin_bp.route('/games/cleanup-expired', methods=['POST'])
@login_required
@admin_required
def cleanup_expired_games():
    from app.utils.common import clean_expired_rooms
    from datetime import datetime as dt
    before_waiting = GameRoom.query.filter_by(status='waiting').count()
    before_playing = GameRoom.query.filter_by(status='playing').count()
    
    clean_expired_rooms()
    
    after_waiting = GameRoom.query.filter_by(status='waiting').count()
    after_playing = GameRoom.query.filter_by(status='playing').count()
    
    return jsonify({
        'success': True,
        'cleaned_waiting': before_waiting - after_waiting,
        'cleaned_playing': before_playing - after_playing,
        'message': f'已清理 {before_waiting - after_waiting} 个等待中房间, {before_playing - after_playing} 个进行中超时房间'
    })

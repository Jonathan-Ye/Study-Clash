"""动态生成题目导入Excel模板

含学科和章节下拉选择、填写说明工作表。
学科与章节列支持级联联动：选择学科后，章节下拉自动过滤为该学科下的章节。
通过隐藏数据工作表 + Named Range + INDIRECT函数实现。
"""

import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from app.models import Subject, Chapter


def _sanitize_named_range_name(name: str) -> str:
    """将学科名称转换为合法的Excel Defined Name名称

    Excel Defined Name规则：
    - 首字符必须为字母或下划线
    - 仅允许字母、数字、下划线、点号（Excel支持Unicode字符，中文合法）
    - 最大255字符
    - 不能与单元格引用冲突（如A1, R1C1）

    Args:
        name: 原始学科名称

    Returns:
        str: 合法化的Named Range名称
    """
    if not name:
        return '_blank'

    # 首字符为数字时添加前缀
    if name[0].isdigit():
        name = '_' + name

    # 替换非法字符为下划线（\w包含Unicode字母、数字、下划线）
    name = re.sub(r'[^\w.]', '_', name)

    # 检查是否与Excel单元格引用冲突
    if re.match(r'^[A-Z]+[0-9]+$', name) or re.match(r'^R[0-9]+C[0-9]+$', name):
        name = '_nr_' + name

    # 截断至255字符
    return name[:255]


def _sort_chapters_tree(chapters):
    """将章节列表按树形结构排序：父章节在前，子章节紧跟其后

    排序规则与前端章节管理页面的拖动排序一致：
    1. 顶级章节按order排序
    2. 每个父章节后面紧跟其子章节（子章节也按order排序）
    3. 递归处理更深层级

    Args:
        chapters: Chapter对象列表（同一学科下）

    Returns:
        list: 按树形结构排序后的Chapter列表
    """
    # 按parent_id分组
    children_map = {}  # parent_id -> [child_chapters]
    root_chapters = []
    for c in chapters:
        if c.parent_id is None:
            root_chapters.append(c)
        else:
            children_map.setdefault(c.parent_id, []).append(c)

    # 每组内按order排序
    root_chapters.sort(key=lambda x: x.order or 0)
    for pid in children_map:
        children_map[pid].sort(key=lambda x: x.order or 0)

    # 递归构建排序结果
    result = []
    def add_with_children(parent):
        result.append(parent)
        for child in children_map.get(parent.id, []):
            add_with_children(child)

    for root in root_chapters:
        add_with_children(root)

    return result


def _create_chapter_data_sheet(wb, subjects, subject_chapters):
    """创建隐藏数据工作表，存储各学科章节选项，并定义Named Range

    在工作簿中创建一个隐藏的"章节选项"工作表，按学科分列存储章节路径，
    并为每个学科创建一个Defined Name（Named Range），引用该学科对应的
    章节路径单元格区域。INDIRECT函数通过该Named Range实现级联下拉。

    Args:
        wb: openpyxl Workbook对象
        subjects: Subject对象列表
        subject_chapters: {subject.id: [chapter_full_path, ...]}
            学科ID到章节路径列表的映射，路径不含学科前缀

    Returns:
        dict: {subject.name: named_range_name}
            学科名称到Named Range名称的映射，用于INDIRECT引用时
            查找合法化的Named Range名称
    """
    ws_source = wb.create_sheet(title="章节选项")
    name_mapping = {}

    for col_idx, subject in enumerate(subjects, 1):
        # 获取合法的Named Range名称
        nr_name = _sanitize_named_range_name(subject.name)
        name_mapping[subject.name] = nr_name

        col_letter = get_column_letter(col_idx)

        # 列标题写入学科名称（便于用户取消隐藏后理解数据结构）
        ws_source.cell(row=1, column=col_idx, value=subject.name)

        # 获取该学科的章节路径列表
        chapters = subject_chapters.get(subject.id, [])

        # 从第2行开始写入章节路径
        for row_idx, chapter_path in enumerate(chapters, 2):
            ws_source.cell(row=row_idx, column=col_idx, value=chapter_path)

        # 创建Defined Name（Named Range）
        if chapters:
            # 有章节：引用第2行到最后一行数据
            last_row = len(chapters) + 1
            ref = f"'{ws_source.title}'!${col_letter}$2:${col_letter}${last_row}"
        else:
            # 无章节：引用空区域（列标题行自身，下拉为空）
            ref = f"'{ws_source.title}'!${col_letter}$1:${col_letter}$1"

        dn = DefinedName(name=nr_name, attr_text=ref)
        wb.defined_names.add(dn)

    # 隐藏工作表（用户可通过右键标签取消隐藏查看）
    ws_source.sheet_state = 'hidden'

    return name_mapping


def _create_cascade_validation(ws_data, subject_names, name_mapping):
    """为学科列和章节列创建级联DataValidation

    学科列（M列）使用普通下拉列表，章节列（N列）使用INDIRECT函数
    引用学科列的值，实现级联下拉。当用户选择学科后，INDIRECT函数
    自动查找该学科名称对应的Named Range，从而显示该学科下的章节。

    INDIRECT(M2)是相对引用，Excel会自动为每行偏移：
    N3引用M3，N4引用M4...，无需逐行设置不同公式。

    Args:
        ws_data: 题目数据工作表
        subject_names: 学科名称列表
        name_mapping: {subject.name: named_range_name}
            学科名称到Named Range名称的映射
    """
    if not subject_names:
        return

    # ===== 学科列DataValidation（M列）=====
    # 构建学科选项：若名称经合法化后不同，使用合法化名称（确保INDIRECT能找到）
    # 若合法化后与原始名相同，直接使用原始名
    display_names = []
    for name in subject_names:
        nr_name = name_mapping.get(name, name)
        display_names.append(nr_name if nr_name != name else name)

    subject_options = ','.join(display_names)
    # Excel DataValidation formula1 限制255字符
    if len(subject_options) > 255:
        truncated = []
        total_len = 0
        for name in display_names:
            item = name if not truncated else ',' + name
            if total_len + len(item) > 255:
                break
            truncated.append(name)
            total_len += len(item)
        subject_options = ','.join(truncated)

    subject_dv = DataValidation(
        type="list",
        formula1=f'"{subject_options}"',
        allow_blank=True
    )
    subject_dv.error = "请从下拉列表中选择学科"
    subject_dv.errorTitle = "学科选择"
    ws_data.add_data_validation(subject_dv)
    subject_dv.add('M2:M1000')

    # ===== 章节列DataValidation（N列）=====
    # 使用INDIRECT函数引用学科列的值，实现级联下拉
    chapter_dv = DataValidation(
        type="list",
        formula1="=INDIRECT(M2)",
        allow_blank=True
    )
    chapter_dv.error = "请先选择学科，再从下拉列表选择章节，或手动输入章节路径"
    chapter_dv.errorTitle = "章节选择"
    chapter_dv.showErrorMessage = False  # 允许手动输入
    ws_data.add_data_validation(chapter_dv)
    chapter_dv.add('N2:N1000')


def generate_import_template():
    """动态生成题目导入Excel模板

    Returns:
        BytesIO: 包含Excel模板数据的字节流
    """
    wb = Workbook()

    # ===== 题目数据工作表 =====
    ws_data = wb.active
    ws_data.title = "题目数据"

    headers = [
        'ID', '题目内容', '题型', '难度', '正确答案',
        '选项A', '选项B', '选项C', '选项D', '选项E', '选项F',
        '解析', '学科', '章节'
    ]

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 示例数据（章节列不含学科前缀，学科信息由学科列单独提供）
    sample_data = [
        [None, '1+1等于多少？', '单选', '简单', 'A',
         '1', '2', '3', '4', '', '',
         '1+1=2', '数学', '第一章 > 基础运算'],
        [None, '中国的首都是哪里？', '单选', '简单', 'B',
         '上海', '北京', '广州', '深圳', '', '',
         '北京是中国的首都', '语文', '第一章'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 设置列宽
    col_widths = {
        'A': 8, 'B': 30, 'C': 10, 'D': 10, 'E': 12,
        'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 10, 'K': 10,
        'L': 20, 'M': 12, 'N': 25
    }
    for col_letter, width in col_widths.items():
        ws_data.column_dimensions[col_letter].width = width

    ws_data.row_dimensions[1].height = 30

    # ===== 数据验证：学科-章节级联下拉 =====
    subjects = Subject.query.filter_by(is_active=True).all()
    subject_names = [s.name for s in subjects]

    # 按学科分组查询章节，构建学科ID到章节路径列表的映射
    # 按树形结构排序，与前端章节管理页面的拖动排序一致
    subject_chapters = {}
    for subject in subjects:
        chapters = Chapter.query.filter_by(
            subject_id=subject.id, is_active=True
        ).all()
        sorted_chapters = _sort_chapters_tree(chapters)
        subject_chapters[subject.id] = [c.get_full_path() for c in sorted_chapters]

    # 创建隐藏数据工作表和Named Range
    name_mapping = _create_chapter_data_sheet(wb, subjects, subject_chapters)

    # 创建级联DataValidation（学科列普通下拉 + 章节列INDIRECT级联）
    _create_cascade_validation(ws_data, subject_names, name_mapping)

    # ===== 填写说明工作表 =====
    ws_guide = wb.create_sheet(title="填写说明")

    guide_content = [
        ["【必填字段】（3项）", ""],
        ["题目内容", "题目的正文内容"],
        ["题型", "单选 / 多选 / 判断 / 填空"],
        ["正确答案", "如 A 或 A,B,C（多选用逗号分隔）"],
        ["", ""],
        ["【选填字段】", ""],
        ["ID", "题目ID，用于冲突检测（一般留空）"],
        ["难度", "简单 / 中等 / 困难 / 极难，或 1/2/3/4"],
        ["选项A~F", "选择题选项文本，单选/多选至少填A和B"],
        ["解析", "答案解析说明"],
        ["学科", "从下拉列表选择学科名称，选择后章节列下拉将自动显示该学科下的章节，不存在则自动创建"],
        ["章节", "先选择学科，再从下拉列表选择章节（自动过滤），也可手动输入章节路径"],
        ["", ""],
        ["【章节路径格式】", ""],
        ["1级（章）", "如：第一章"],
        ["2级（章 > 节）", "如：第一章 > 第一节"],
        ["3级（章 > 节 > 小节）", "如：第一章 > 第一节 > 第一小节"],
        ["", '使用 > 或 / 分隔层级，如"第一章/第一节"等同于"第一章 > 第一节"'],
        ["", "分隔符前后空格可选，最多支持3级章节层级，超过3级将校验失败"],
        ["", "章节不存在时将自动创建（需同时指定学科）"],
        ["", '也可带学科前缀输入如"数学 > 第一章 > 第一节"，此时以章节路径中的学科名为准'],
        ["", ""],
        ["【注意事项】", ""],
        ["1.", "每行代表一道题目，第一行为表头（不要删除）"],
        ["2.", "请先选择学科，再选择章节，章节下拉会根据学科列的值自动过滤"],
        ["3.", "学科列和章节列也支持手动输入"],
        ["4.", "指定章节时必须同时指定学科，否则该行校验失败"],
        ["5.", "填写完成后保存为 .xlsx 或 .xls 格式即可导入"],
    ]

    title_font = Font(bold=True, size=12, color="4472C4")
    normal_font = Font(size=10)

    for row_idx, (field, desc) in enumerate(guide_content, 1):
        cell_field = ws_guide.cell(row=row_idx, column=1, value=field)
        cell_desc = ws_guide.cell(row=row_idx, column=2, value=desc)

        if field.startswith("【"):
            cell_field.font = title_font
        else:
            cell_field.font = normal_font
        cell_desc.font = normal_font

    ws_guide.column_dimensions['A'].width = 25
    ws_guide.column_dimensions['B'].width = 60

    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

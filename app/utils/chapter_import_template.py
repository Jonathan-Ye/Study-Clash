"""动态生成章节导入Excel模板

含学科下拉选择、层级选择、填写说明工作表。
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from app.models import Subject


def generate_chapter_import_template():
    """动态生成章节导入Excel模板

    Returns:
        BytesIO: 包含Excel模板数据的字节流
    """
    wb = Workbook()

    # ===== 章节数据工作表 =====
    ws_data = wb.active
    ws_data.title = "章节数据"

    headers = ['学科名称', '章节名称', '层级', '父章节名称', '排序', '描述', '启用']

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 示例数据
    sample_data = [
        ['数学', '第一章', 1, '', 1, '', '是'],
        ['数学', '第一节', 2, '第一章', 1, '', '是'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 设置列宽
    col_widths = {'A': 15, 'B': 20, 'C': 8, 'D': 20, 'E': 8, 'F': 30, 'G': 8}
    for col_letter, width in col_widths.items():
        ws_data.column_dimensions[col_letter].width = width

    ws_data.row_dimensions[1].height = 30

    # ===== 数据验证下拉列表 =====
    # 学科名称列（A列）
    subjects = Subject.query.filter_by(is_active=True).all()
    subject_names = [s.name for s in subjects]

    if subject_names:
        subject_options = ','.join(subject_names)
        # Excel DataValidation formula1 限制255字符
        if len(subject_options) > 255:
            truncated = []
            total_len = 0
            for name in subject_names:
                item = name if not truncated else ',' + name
                if total_len + len(item) > 255:
                    break
                truncated.append(name)
                total_len += len(item)
            subject_options = ','.join(truncated)

        subject_dv = DataValidation(
            type="list",
            formula1=f'"{subject_options}"',
            allow_blank=True
        )
        subject_dv.error = "请从下拉列表中选择学科"
        subject_dv.errorTitle = "学科选择"
        ws_data.add_data_validation(subject_dv)
        subject_dv.add('A2:A1000')

    # 层级列（C列）
    level_dv = DataValidation(
        type="list",
        formula1='"1,2,3"',
        allow_blank=True
    )
    level_dv.error = "层级必须为1、2或3"
    level_dv.errorTitle = "层级选择"
    ws_data.add_data_validation(level_dv)
    level_dv.add('C2:C1000')

    # 启用列（G列）
    active_dv = DataValidation(
        type="list",
        formula1='"是,否"',
        allow_blank=True
    )
    active_dv.error = "请选择是或否"
    active_dv.errorTitle = "启用选择"
    ws_data.add_data_validation(active_dv)
    active_dv.add('G2:G1000')

    # ===== 填写说明工作表 =====
    ws_guide = wb.create_sheet(title="填写说明")

    guide_content = [
        ["【必填字段】（2项）", ""],
        ["学科名称", "章节所属学科，从下拉列表选择或手动输入，不存在则自动创建"],
        ["章节名称", '章节的名称，如"第一章"、"第一节"'],
        ["", ""],
        ["【选填字段】", ""],
        ["层级", "1=章，2=节，3=小节，从下拉列表选择，默认为1"],
        ["父章节名称", "上级章节的名称，层级1时留空，层级2时填写章名，层级3时填写节名"],
        ["排序", "同级章节的显示顺序，数字越小越靠前，留空则自动排序"],
        ["描述", "章节的描述说明"],
        ["启用", '是/否，默认为"是"'],
        ["", ""],
        ["【章节层级说明】", ""],
        ["层级1（章）", '顶级章节，如"第一章"、"第二单元"，无需填写父章节名称'],
        ["层级2（节）", '章下的子章节，如"第一节"、"第二课"，需填写父章节名称（章名）'],
        ["层级3（小节）", '节下的子章节，如"第一小节"、"知识点1"，需填写父章节名称（节名）'],
        ["", ""],
        ["【注意事项】", ""],
        ["1.", "每行代表一个章节，第一行为表头（不要删除）"],
        ["2.", "同一学科下同一父章节下不应有同名章节"],
        ["3.", "层级1的章节父章节名称必须留空"],
        ["4.", "层级2和3的章节必须填写父章节名称"],
        ["5.", "导入时按层级升序处理（先导入层级1，再层级2，再层级3）"],
        ["6.", "填写完成后保存为 .xlsx 或 .xls 格式即可导入"],
    ]

    title_font = Font(bold=True, size=12, color="4472C4")
    normal_font = Font(size=10)

    for row_idx, (field, desc) in enumerate(guide_content, 1):
        cell_field = ws_guide.cell(row=row_idx, column=1, value=field)
        cell_desc = ws_guide.cell(row=row_idx, column=2, value=desc)

        if field.startswith("【"):
            cell_field.font = title_font
        else:
            cell_field.font = normal_font
        cell_desc.font = normal_font

    ws_guide.column_dimensions['A'].width = 25
    ws_guide.column_dimensions['B'].width = 60

    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
